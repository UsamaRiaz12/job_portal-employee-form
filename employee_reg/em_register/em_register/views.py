from django.http import JsonResponse
from django.shortcuts import redirect, render
from .models import em_register
from .serializer import em_registerSerilaizer
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status 
from django.shortcuts import render, redirect
from em_register.models import em_register
from rest_framework.views import APIView
import openpyxl

# def em_register_view(request):
#         objects = em_register.objects.all() 


@api_view(['GET', 'POST'])
def em_register_list(request,format=None):
    if request.method == 'GET':
        em_registers = em_register.objects.all()
        serializer = em_registerSerilaizer(em_registers, many=True)
        return JsonResponse({'em_register': serializer.data})

    elif request.method == 'POST':
        serializer = em_registerSerilaizer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT', 'DELETE'])
def em_register_detail(request, id,format=None):
    try:
        em_reg = em_register.objects.get(pk=id)
    except em_register.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
         
    if request.method == 'GET':
        serializer = em_registerSerilaizer(em_reg)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = em_registerSerilaizer(em_reg, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == 'DELETE':
        em_reg.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)


def em_register_view(request):
    if request.method == 'POST':
        form = em_registerSerilaizer(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('success_page')  # Redirect to a success page
    else:
        form = em_registerSerilaizer()
    return render(request, 'em_register.py', {'form': form})





# class xport_selected_to_excel(APIView):
#     def post(self, request, *args, **kwargs):
#         queryset = em_register.objects.all()

#         workbook = openpyxl.Workbook()
#         worksheet = workbook.active

#         # Write headers
#         headers = ['Name', 'Email', 'City', 'Number']
#         for col_num, header_title in enumerate(headers, 1):
#             worksheet.cell(row=1, column=col_num, value=header_title)

#         # Write data
#         for row_num, item in enumerate(queryset, 2):
#             worksheet.cell(row=row_num, column=1, value=item.name)
#             worksheet.cell(row=row_num, column=2, value=item.email)
#             worksheet.cell(row=row_num, column=3, value=item.city)
#             worksheet.cell(row=row_num, column=4, value=item.number)

#         # Save the workbook
#         file_path = 'exported_data.xlsx'
#         workbook.save(file_path)

#         return Response({'message': 'Data exported to Excel.'})
